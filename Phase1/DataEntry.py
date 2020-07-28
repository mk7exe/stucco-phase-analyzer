import openpyxl
from PyQt5 import QtWidgets as qtw
from PyQt5 import QtCore as qtc
from PyQt5 import QtGui as qtg

class FloatDelegate(qtw.QStyledItemDelegate):
    def __init__(self, decimals, parent=None):
        super(FloatDelegate, self).__init__(parent=parent)
        self.nDecimals = decimals

    def displayText(self, value, locale):
        try:
            number = float(value)
        except ValueError:
            return super(FloatDelegate, self).displayText(value, locale)
        else:
            return locale.toString(number, format="f", precision=self.nDecimals)

    def createEditor(self, parent, option, index):
        lineEdit = qtw.QLineEdit(parent)
        lineEdit.setValidator(qtg.QDoubleValidator())
        return lineEdit


class ReadOnlyDelegate(qtw.QStyledItemDelegate):

    def createEditor(self, parent, option, index):
        return


class MyTableReference(qtw.QTableWidget):

    changed = qtc.pyqtSignal(int, int)

    def __init__(self, row, col):
        super().__init__(row, col)
        self.setHorizontalHeaderLabels(['Label', 'Sample 1', 'Sample 2', 'Sample 3'])
        self.verticalHeader().setDefaultSectionSize(25)
        self.horizontalHeader().setDefaultSectionSize(400)
        self.horizontalHeader().setSectionResizeMode(qtw.QHeaderView.Fixed)
        stylesheet = "QHeaderView::section{border-radius: 14px;}"
        self.setStyleSheet(stylesheet)

    def _addRow(self):
        self.changed.emit(self.rowCount(), 0)
        self.insertRow(self.rowCount())

    def _removeRow(self):
        if self.rowCount() > 1:
            self.changed.emit(self.currentRow(), 1)
            self.removeRow(self.currentRow())

    def keyPressEvent(self, event):
        key = event.key()

        if key == qtc.Qt.Key_Return or key == qtc.Qt.Key_Enter:
            self._addRow()
        else:
            super(MyTableReference, self).keyPressEvent(event)


class MyTable(qtw.QTableWidget):

    def __init__(self, row, col):
        super().__init__(row, col)
        self.setHorizontalHeaderLabels(['Label', 'Sample 1', 'Sample 2', 'Sample 3'])
        self.verticalHeader().setDefaultSectionSize(25)
        self.horizontalHeader().setDefaultSectionSize(120)
        self.horizontalHeader().setSectionResizeMode(qtw.QHeaderView.Fixed)
        stylesheet = "QHeaderView::section{border-radius: 14px;}"
        self.setStyleSheet(stylesheet)
        self.setColumnWidth(0, 400)

        # first column is read only and is updated from tab 1
        delegate1 = ReadOnlyDelegate(self)
        self.setItemDelegateForColumn(0, delegate1)

        # only accept numbers for comulmn 2-4 and show numbers up to 4 decimal numbers
        delegate2 = FloatDelegate(4)
        self.setItemDelegate(delegate2)

    def _update(self, row, signal):
        if signal == 0:
            self.insertRow(row)
        if signal == 1:
            self.removeRow(row)


class MyTabWidget(qtw.QWidget):

    done = qtc.pyqtSignal(int)

    def __init__(self, samples):
        super(qtw.QWidget, self).__init__()

        self.samples = samples
        self.layout = qtw.QVBoxLayout(self)

        # Initializing tabs
        self.tabs = qtw.QTabWidget()
        self.tab1 = qtw.QWidget()
        self.tab2 = qtw.QWidget()
        self.tab3 = qtw.QWidget()
        self.tab4 = qtw.QWidget()

        # Adding tabs
        self.tabs.addTab(self.tab1, 'Samples')
        self.tabs.addTab(self.tab2, 'ORG')
        self.tabs.addTab(self.tab3, 'HUM')
        self.tabs.addTab(self.tab4, 'HYD')

        # Creating tab1
        self.tab1.layout = qtw.QHBoxLayout(self)
        # tabel
        self.tab1.table = MyTableReference(len(self.samples) + 1, 1)
        self.tab1.layout.addWidget(self.tab1.table)
        # push buttons
        self.tab1.buttonlayout = qtw.QVBoxLayout(self)
        self.pushButtonEnter = qtw.QPushButton("Add")
        self.pushButtonEnter.clicked.connect(self.tab1.table._addRow)
        self.pushButtonDelete = qtw.QPushButton("Remove")
        self.pushButtonDelete.clicked.connect(self.tab1.table._removeRow)
        self.pushButtonDone = qtw.QPushButton("Submit")
        self.pushButtonDone.clicked.connect(self._submit)
        self.tab1.buttonlayout.addWidget(self.pushButtonEnter)
        self.tab1.buttonlayout.addWidget(self.pushButtonDelete, alignment=qtc.Qt.AlignTop)
        self.tab1.buttonlayout.addWidget(self.pushButtonDone, alignment=qtc.Qt.AlignBottom)
        # Assign the layout for Tab1
        self.tab1.layout.addLayout(self.tab1.buttonlayout)
        self.tab1.setLayout(self.tab1.layout)

        # Creating tab2
        self.tab2.layout = qtw.QVBoxLayout(self)
        self.tab2.tabs = qtw.QTabWidget()
        self.tab2.tab1 = qtw.QWidget()
        self.tab2.tabs.addTab(self.tab2.tab1, 'Weight loss measurements (%)')
        # tabel
        self.tab2.tab1.layout = qtw.QVBoxLayout(self)
        self.tab2.tab1.table = MyTable(len(self.samples) + 1, 4)
        self.tab2.tab1.table.reference = self.tab1.table
        self.tab2.tab1.table.reference.changed.connect(self.tab2.tab1.table._update)
        self.tab2.tab1.layout.addWidget(self.tab2.tab1.table)
        self.tab2.tab1.setLayout(self.tab2.tab1.layout)
        self.tab2.layout.addWidget(self.tab2.tabs)
        self.tab2.setLayout(self.tab2.layout)

        # Creating tab3
        self.tab3.layout = qtw.QVBoxLayout(self)
        self.tab3.tabs = qtw.QTabWidget()
        self.tab3.tab1 = qtw.QWidget()
        self.tab3.tab2 = qtw.QWidget()
        self.tab3.tabs.addTab(self.tab3.tab1, 'Weight loss measurements (%)')
        self.tab3.tabs.addTab(self.tab3.tab2, 'Weight gain measurements (%)')
        # tabels wl
        self.tab3.tab1.layout = qtw.QVBoxLayout(self)
        self.tab3.tab1.table = MyTable(len(self.samples) + 1, 4)
        self.tab3.tab1.table.reference = self.tab1.table
        self.tab3.tab1.table.reference.changed.connect(self.tab3.tab1.table._update)
        self.tab3.tab1.layout.addWidget(self.tab3.tab1.table)
        self.tab3.tab1.setLayout(self.tab3.tab1.layout)
        # tabels wg
        self.tab3.tab2.layout = qtw.QVBoxLayout(self)
        self.tab3.tab2.table = MyTable(len(self.samples) + 1, 4)
        self.tab3.tab2.table.reference = self.tab1.table
        self.tab3.tab2.table.reference.changed.connect(self.tab3.tab2.table._update)
        self.tab3.tab2.layout.addWidget(self.tab3.tab2.table)
        self.tab3.tab2.setLayout(self.tab3.tab2.layout)

        self.tab3.layout.addWidget(self.tab3.tabs)
        self.tab3.setLayout(self.tab3.layout)

        # Creating tab3
        self.tab4.layout = qtw.QVBoxLayout(self)
        self.tab4.tabs = qtw.QTabWidget()
        self.tab4.tab1 = qtw.QWidget()
        self.tab4.tab2 = qtw.QWidget()
        self.tab4.tabs.addTab(self.tab4.tab1, 'Weight loss measurements (%)')
        self.tab4.tabs.addTab(self.tab4.tab2, 'Weight gain measurements (%)')
        # tabels wl
        self.tab4.tab1.layout = qtw.QVBoxLayout(self)
        self.tab4.tab1.table = MyTable(len(self.samples) + 1, 4)
        self.tab4.tab1.table.reference = self.tab1.table
        self.tab4.tab1.table.reference.changed.connect(self.tab4.tab1.table._update)
        self.tab4.tab1.layout.addWidget(self.tab4.tab1.table)
        self.tab4.tab1.setLayout(self.tab4.tab1.layout)
        # tabels wg
        self.tab4.tab2.layout = qtw.QVBoxLayout(self)
        self.tab4.tab2.table = MyTable(len(self.samples) + 1, 4)
        self.tab4.tab2.table.reference = self.tab1.table
        self.tab4.tab2.table.reference.changed.connect(self.tab4.tab2.table._update)
        self.tab4.tab2.layout.addWidget(self.tab4.tab2.table)
        self.tab4.tab2.setLayout(self.tab4.tab2.layout)

        self.tab4.layout.addWidget(self.tab4.tabs)
        self.tab4.setLayout(self.tab4.layout)

        # Add tabs to widget
        self.layout.addWidget(self.tabs)
        self.setLayout(self.layout)

        # updating labels in all tables according to tab 1
        self.tab1.table.itemChanged.connect(self.tables_update)

        #populate tables with current values
        if len(self.samples) > 0:
            self.tables_populate()

    def tables_populate(self):
        for i in range(len(self.samples)):
            self.tab1.table.setItem(i, 0, qtw.QTableWidgetItem(self.samples[i]['Label']))
            for j in range(3):
                if j < len(self.samples[i]['wl']['org']):
                    self.tab2.tab1.table.setItem(i, j + 1, qtw.QTableWidgetItem(str(self.samples[i]['wl']['org'][j])))
                if j < len(self.samples[i]['wl']['hum']):
                    self.tab3.tab1.table.setItem(i, j + 1, qtw.QTableWidgetItem(str(self.samples[i]['wl']['hum'][j])))
                if j < len(self.samples[i]['wl']['hyd']):
                    self.tab4.tab1.table.setItem(i, j + 1, qtw.QTableWidgetItem(str(self.samples[i]['wl']['hyd'][j])))
                if j < len(self.samples[i]['wg']['hum']):
                    self.tab3.tab2.table.setItem(i, j + 1, qtw.QTableWidgetItem(str(self.samples[i]['wg']['hum'][j])))
                if j < len(self.samples[i]['wg']['hyd']):
                    self.tab4.tab2.table.setItem(i, j + 1, qtw.QTableWidgetItem(str(self.samples[i]['wg']['hyd'][j])))

    def tables_update(self, item):
        i = item.row()
        if not item.text() is None:
            self.tab2.tab1.table.setItem(i, 0, qtw.QTableWidgetItem(item.text()))
            self.tab3.tab1.table.setItem(i, 0, qtw.QTableWidgetItem(item.text()))
            self.tab3.tab2.table.setItem(i, 0, qtw.QTableWidgetItem(item.text()))
            self.tab4.tab1.table.setItem(i, 0, qtw.QTableWidgetItem(item.text()))
            self.tab4.tab2.table.setItem(i, 0, qtw.QTableWidgetItem(item.text()))

    def _submit(self):
        tableLabel = self.tab1.table
        tablewlorg = self.tab2.tab1.table
        tablewlhum = self.tab3.tab1.table
        tablewghum = self.tab3.tab2.table
        tablewlhyd = self.tab4.tab1.table
        tablewghyd = self.tab4.tab2.table

        for row in range(tableLabel.rowCount()):
            column = 0
            _item = tableLabel.item(row, column)
            if _item:
                item = tableLabel.item(row, column).text()
                self.samples[row] = {}
                self.samples[row]['Label'] = item
                self.samples[row]['wl'] = {'org': [], 'hum': [], 'hyd': []}
                self.samples[row]['wg'] = {'hum': [], 'hyd': []}
                for column in range(3):
                    _item = tablewlorg.item(row, column + 1)
                    if _item:
                        item = tablewlorg.item(row, column + 1).text()
                        self.samples[row]['wl']['org'].append(float(item))
                    _item = tablewlhum.item(row, column + 1)
                    if _item:
                        item = tablewlhum.item(row, column + 1).text()
                        self.samples[row]['wl']['hum'].append(float(item))
                    _item = tablewlhyd.item(row, column + 1)
                    if _item:
                        item = tablewlhyd.item(row, column + 1).text()
                        self.samples[row]['wl']['hyd'].append(float(item))
                    _item = tablewghum.item(row, column + 1)
                    if _item:
                        item = tablewghum.item(row, column + 1).text()
                        self.samples[row]['wg']['hum'].append(float(item))
                    _item = tablewghyd.item(row, column + 1)
                    if _item:
                        item = tablewghyd.item(row, column + 1).text()
                        self.samples[row]['wg']['hyd'].append(float(item))

        self.done.emit(1)
        return self.samples


class DataEntryWin(qtw.QMainWindow):

    submitted = qtc.pyqtSignal(int)

    def __init__(self, samples):
        super().__init__()
        self.setWindowModality(qtc.Qt.ApplicationModal)
        self.samples = samples
        self.ini_samples = samples
        self.title = 'Data Entry'
        self.left = 100
        self.top = 100
        self.width = 900
        self.height = 600
        self.setWindowTitle(self.title)
        self.setGeometry(self.left, self.top, self.width, self.height)

        self.menuBar = self.menuBar()
        self.fileMenu = self.menuBar.addMenu('File')

        # export to Excel file action
        exportAction = qtw.QAction('Export to Excel', self)
        exportAction.setShortcut('Ctrl+E')
        exportAction.triggered.connect(self.export_to_excel)

        # exit action
        exitAction = qtw.QAction('Exit', self)
        exitAction.setShortcut('Ctrl+Q')
        exitAction.triggered.connect(lambda: self.close())

        self.fileMenu.addAction(exportAction)
        self.fileMenu.addAction(exitAction)

        self.w = MyTabWidget(self.samples)
        self.setCentralWidget(self.w)
        self.data_submitted = False
        # self.w.done.connect(self.submit_happened)
        self.w.pushButtonDone.clicked.connect(self.submit_happened)

    def submit_happened(self):
        self.samples = self.w._submit()
        msgBox = qtw.QMessageBox()
        msgBox.setIcon(qtw.QMessageBox.Information)
        msgBox.setText(str(len(self.samples))+" samples were submitted.")
        msgBox.setStandardButtons(qtw.QMessageBox.Ok)
        returnValue = msgBox.exec()
        self.data_submitted = True
        self.submitted.emit(1)

    def export_to_excel(self):
        if self.data_submitted:
            fileopen = qtw.QFileDialog.getSaveFileName(self, "Save File", filter="Excel files (*.xlsx *.xls)")
            filename = fileopen[0]
            wb = openpyxl.Workbook()
            ws0 = wb.active
            ws0.title = "Labels"
            ws1 = wb.create_sheet("Weight Loss")
            ws2 = wb.create_sheet("Weight Gain")
            thin = openpyxl.styles.Side(border_style="thin")
            double = openpyxl.styles.Side(border_style="double")
            color = ['DDDDDD', 'DDDDDD', 'DDDDDD', 'C7CEFF', 'C7CEFF', 'C7CEFF', 'FFC7CE', 'FFC7CE', 'FFC7CE']
            ws_count = 0
            for ws in wb:
                # stylings
                ws.column_dimensions["A"].width = 35
                ws.cell(row=1, column=1, value='Labels')
                if ws_count > 0:
                    ws.merge_cells('B1:D1')
                    ws['B1'] = 'ORG'
                    ws.merge_cells('E1:G1')
                    ws['E1'] = 'HUM'
                    ws.merge_cells('H1:J1')
                    ws['H1'] = 'HYD'
                    for i in range(100):
                        for j in range(9):
                            c = ws.cell(row=i + 1, column=j + 2)
                            c.fill = openpyxl.styles.PatternFill("solid", fgColor=color[j])
                            c.alignment = openpyxl.styles.Alignment(horizontal='center')
                            if j % 3 == 1:
                                c.border = openpyxl.styles.Border(top=thin, left=thin, right=thin, bottom=thin)
                            if j % 3 == 0:
                                c.border = openpyxl.styles.Border(top=thin, left=double, right=thin, bottom=thin)
                            if j % 3 == 2:
                                c.border = openpyxl.styles.Border(top=thin, left=thin, right=double, bottom=thin)
                else:
                    for i in range(100):
                        c = ws.cell(row=i + 1, column=1)
                        c.alignment = openpyxl.styles.Alignment(horizontal='left')
                        c.border = openpyxl.styles.Border(top=thin, left=double, right=double, bottom=thin)

                for c in ws[1]:
                    c.font = openpyxl.styles.Font(bold=True)
                ws_count += 1

            for i in range(len(self.samples)):
                ws0.cell(row=i+2, column=1, value=self.samples[i]['Label'])
                ws1.cell(row=i + 2, column=1, value=self.samples[i]['Label'])
                ws2.cell(row=i + 2, column=1, value=self.samples[i]['Label'])
                for j in range(3):
                    if j < len(self.samples[i]['wl']['org']):
                        ws1.cell(row=i + 2, column=j + 2, value=self.samples[i]['wl']['org'][j])
                        # ws1.cell(row=i + 2, column=j + 2).alignment = openpyxl.styles.Alignment(horizontal='center')
                    if j < len(self.samples[i]['wl']['hum']):
                        ws1.cell(row=i + 2, column=j + 5, value=self.samples[i]['wl']['hum'][j])
                        # ws1.cell(row=i + 2, column=j + 5).alignment = openpyxl.styles.Alignment(horizontal='center')
                    if j < len(self.samples[i]['wl']['hyd']):
                        ws1.cell(row=i + 2, column=j + 8, value=self.samples[i]['wl']['hyd'][j])
                        # ws1.cell(row=i + 2, column=j + 8).alignment = openpyxl.styles.Alignment(horizontal='center')
                    if j < len(self.samples[i]['wg']['hum']):
                        ws2.cell(row=i + 2, column=j + 5, value=self.samples[i]['wg']['hum'][j])
                        # ws2.cell(row=i + 2, column=j + 5).alignment = openpyxl.styles.Alignment(horizontal='center')
                    if j < len(self.samples[i]['wg']['hyd']):
                        ws2.cell(row=i + 2, column=j + 8, value=self.samples[i]['wg']['hyd'][j])
                        # ws2.cell(row=i + 2, column=j + 8).alignment = openpyxl.styles.Alignment(horizontal='center')


            wb.save(filename)
        else:
            msgBox = qtw.QMessageBox()
            msgBox.setIcon(qtw.QMessageBox.Critical)
            msgBox.setText("No sample was submitted. Nothing to save!")
            msgBox.setStandardButtons(qtw.QMessageBox.Ok)
            returnValue = msgBox.exec()

    def closeEvent(self, event):
        if not self.data_submitted and len(self.samples) > 0 and self.ini_samples != self.samples:
            reply = qtw.QMessageBox.critical(self, 'Window Close', 'You have not submitted your data entry. \n' +
                                             'Your data may be lost if you close the window.\n' +
                                             'Are you sure you want to close the window?',
                                             qtw.QMessageBox.Yes | qtw.QMessageBox.No, qtw.QMessageBox.No)
            if reply == qtw.QMessageBox.Yes:
                event.accept()
            else:
                event.ignore()
        else:
            event.accept()


# if __name__ == '__main__':
#     app = qtw.QApplication(sys.argv)
#     app.setStyleSheet('QPushButton{font-size: 18px; width: 200px; height: 50px}')
#     samples = {}
#     ex = DataEntryWin(samples)
#     ex.show()
#     # ex.tab_widget.tab1.table.setItem(0, 0, qtw.QTableWidgetItem('test'))
#     sys.exit(app.exec_())
