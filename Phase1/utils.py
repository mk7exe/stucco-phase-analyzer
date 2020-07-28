import openpyxl
import numpy as np
from scipy.optimize import lsq_linear

w0 = 136.14

DW = {
    "dhwl": 2 * 18.01528 / (w0 + 2 * 18.01528),
    "hhwl": 0.5 * 18.01528 / (w0 + 0.5 * 18.01528),
    "hhwg": 1.5 * 18.01528 / (w0 + 0.5 * 18.01528),
    "a3wg1": 0.5 * 18.01528 / w0,
    "a3wg2": 2 * 18.01528 / w0
}


def isdigit(str0):
    str1 = str0.replace('.', '', 1)
    str2 = str1.replace('-', '', 1)
    str3 = str2.replace('+', '', 1)
    return str3.isdigit()


def isListEmpty(inList):
    if isinstance(inList, list):  # Is a list
        return all(map(isListEmpty, inList))
    return False  # Not a list


def gpa_linear(wl_id, dwl, dwa, dwh):
    """
    Function to calculate components of system of linear equations giving stucco phase content from two weight gain
    and one weight loss measurements.
    :param wl_id: The id of the weight loss measurement. 1: ORG sample, 2: HUM sample, 3: HYD sample
    :param dwl: weight loss wl_id sample
    :param dwa: weight gain of HUM sample
    :param dwh: weight gain of HYD sample
    :return:
    A: 4x4 matrix containing coefficients of system of linear equations giving X=[DH, HH, AIII, FM, IN]
    b: array of length 4 giving the constants in AX - b = 0 equation.
    """
    A = np.zeros((4, 4))
    b = np.zeros(4)
    if dwa >= 0:
        A[0, 0] = 0
        A[0, 1] = 0
        A[0, 2] = DW["a3wg1"]
        A[0, 3] = 0
        A[1, 0] = 0
        A[1, 1] = DW["hhwg"]
        A[1, 2] = DW["a3wg2"]
        A[1, 3] = 0
        A[3, 0] = 1
        A[3, 1] = 1
        A[3, 2] = 1
        A[3, 3] = 1
        if wl_id == 0:
            A[2, 0] = DW["dhwl"]
            A[2, 1] = DW["hhwl"]
            A[2, 2] = 0
            A[2, 3] = 0
            b[2] = dwl
        if wl_id == 1:
            A[2, 0] = DW["dhwl"]
            A[2, 1] = DW["hhwl"]
            A[2, 2] = (1 + DW["a3wg1"]) * DW["hhwl"] - DW["a3wg1"] * dwl / 100
            A[2, 3] = 0
            b[2] = dwl
        if wl_id == 2:
            A[2, 0] = DW["dhwl"]
            A[2, 1] = (1 + DW["hhwg"]) * DW["dhwl"] - DW["hhwg"] * dwl / 100
            A[2, 2] = (1 + DW["a3wg2"]) * DW["dhwl"] - DW["a3wg2"] * dwl / 100
            A[2, 3] = 0
            b[2] = dwl
        b[0] = dwa
        b[1] = dwh
        b[3] = 100
    else:
        A[0, 0] = 0
        A[0, 1] = 0
        A[0, 2] = -1
        A[0, 3] = 0
        A[1, 0] = 0
        A[1, 1] = DW["hhwg"]
        A[1, 2] = -1
        A[1, 3] = 0
        A[3, 0] = 1
        A[3, 1] = 1
        A[3, 2] = 1
        A[3, 3] = 1
        if wl_id == 0:
            A[2, 0] = DW["dhwl"]
            A[2, 1] = DW["hhwl"]
            A[2, 2] = 1
            A[2, 3] = 0
            b[2] = dwl
        if wl_id == 1:
            A[2, 0] = DW["dhwl"]
            A[2, 1] = DW["hhwl"]
            A[2, 2] = 0.01 * dwl
            A[2, 3] = 0
            b[2] = dwl
        if wl_id == 2:
            A[2, 0] = DW["dhwl"]
            A[2, 1] = (1 + DW["hhwg"]) * DW["dhwl"] - DW["hhwg"] * dwl / 100
            A[2, 2] = dwl / 100
            A[2, 3] = 0
            b[2] = dwl
        b[0] = dwa
        b[1] = dwh
        b[3] = 100
    return A, b


def tga_linear(dwo, dwa, dwh):
    """
    Function to calculate components of system of linear equations giving stucco phase content from three weight loss
    measurements.
    :param dwo: ORG sample weight loss
    :param dwa: HUM sample weight loss
    :param dwh: HYD sample weight loss
    :return:
    A: 4x4 matrix containing coefficients of system of linear equations giving X=[DH, HH, AIII, FM, IN]
    b: array of length 4 giving the constants in AX - b = 0 equation.
    """
    A = np.zeros((4, 4))
    b = np.zeros(4)
    if dwa >= dwo:
        A[0, 0] = DW["dhwl"]
        A[0, 1] = DW["hhwl"]
        A[0, 2] = 0
        A[0, 3] = 0
        A[1, 0] = DW["dhwl"]
        A[1, 1] = DW["hhwl"]
        A[1, 2] = DW["hhwl"] + DW["hhwl"] * DW["a3wg1"] - 0.01 * DW["a3wg1"] * dwa
        A[1, 3] = 0
        A[2, 0] = DW["dhwl"]
        A[2, 1] = DW["dhwl"] + DW["dhwl"] * DW["hhwg"] - 0.01 * DW["hhwg"] * dwh
        A[2, 2] = DW["dhwl"] + DW["dhwl"] * DW["a3wg2"] - 0.01 * DW["a3wg2"] * dwh
        A[2, 3] = 0
        A[3, 0] = 1
        A[3, 1] = 1
        A[3, 2] = 1
        A[3, 3] = 1
        b[0] = dwo
        b[1] = dwa
        b[2] = dwh
        b[3] = 100
    else:
        A[0, 0] = DW["dhwl"]
        A[0, 1] = DW["hhwl"]
        A[0, 2] = 1
        A[0, 3] = 0
        A[1, 0] = DW["dhwl"]
        A[1, 1] = DW["hhwl"]
        A[1, 2] = 0.01 * DW["a3wg1"] * dwa
        A[1, 3] = 0
        A[2, 0] = DW["dhwl"]
        A[2, 1] = DW["dhwl"] + DW["dhwl"] * DW["hhwg"] - 0.01 * DW["hhwg"] * dwh
        A[2, 2] = 0.01 * DW["a3wg2"] * dwh
        A[2, 3] = 0
        A[3, 0] = 1
        A[3, 1] = 1
        A[3, 2] = 1
        A[3, 3] = 1
        b[0] = dwo
        b[1] = dwa
        b[2] = dwh
        b[3] = 100
    return A, b


def tga_solver(samples):
    """

    """
    labels = [samples[i]['Label'] for i in range(len(samples))]
    wl_org = [samples[i]['wl']['org'] for i in range(len(samples))]
    wl_hum = [samples[i]['wl']['hum'] for i in range(len(samples))]
    wl_hyd = [samples[i]['wl']['hyd'] for i in range(len(samples))]

    tga_pa = [[] for i in range(len(labels))]
    lbond = np.zeros(4)
    ubond = lbond + 100

    for i1 in range(len(labels)):
        if labels[i1]:
            for i2 in range(len(wl_org[i1])):
                if isdigit(str(wl_org[i1][i2])):
                    for i3 in range(len(wl_hum[i1])):
                        if isdigit(str(wl_hum[i1][i3])):
                            for i4 in range(len(wl_hyd[i1])):
                                if isdigit(str(wl_hyd[i1][i4])) and i2 == i3 and i2 == i4:
                                    A, b = tga_linear(wl_org[i1][i2], wl_hum[i1][i3], wl_hyd[i1][i4])
                                    # p = np.linalg.solve(A, b)
                                    lsd = lsq_linear(A, b, bounds=(lbond, ubond), verbose=0)
                                    p = list(lsd.x)
                                    if wl_hum[i1][i3] >= wl_org[i1][i2]:
                                        p.insert(3, 0)
                                    else:
                                        p.insert(2, 0)
                                    tga_pa[i1].append(p)
    tga_pa = [np.array(x) for x in tga_pa]

    phases = {}
    for i in range(len(labels)):
        phases[labels[i]] = dict(zip(["mean", "std"], [np.mean(tga_pa[i], axis=0).tolist(),
                                                       np.std(tga_pa[i], axis=0).tolist()]))

    return phases


def gpa_solver(samples):
    """

    """
    labels = [samples[i]['Label'] for i in range(len(samples))]
    wg_hum = [samples[i]['wg']['hum'] for i in range(len(samples))]
    wg_hyd = [samples[i]['wg']['hyd'] for i in range(len(samples))]

    gpa_pa_org = [[] for i in range(len(labels))]
    gpa_pa_hum = [[] for i in range(len(labels))]
    gpa_pa_hyd = [[] for i in range(len(labels))]

    missmatches = []
    lbond = np.zeros(4)
    ubond = lbond + 100
    tga_wl_id = 0
    for tga_wl_type in ['org', 'hum', 'hyd']:
        for i1 in range(len(labels)):
            if labels[i1]:
                tga_wl = samples[i1]['wl'][tga_wl_type]
                for i2 in range(len(tga_wl)):
                    if isdigit(str(tga_wl[i2])):
                        for i3 in range(len(wg_hum[i1])):
                            if isdigit(str(wg_hum[i1][i3])):
                                for i4 in range(len(wg_hyd[i1])):
                                    if isdigit(str(wg_hyd[i1][i4])) and i2 == i3 and i2 == i4:
                                        A, b = gpa_linear(tga_wl_id, tga_wl[i2], wg_hum[i1][i3], wg_hyd[i1][i4])
                                        # p = list(np.linalg.solve(A, b))
                                        lsd = lsq_linear(A, b, bounds=(lbond, ubond), lsmr_tol='auto', verbose=0)
                                        p = list(lsd.x)
                                        if wg_hum[i1][i3] >= 0:
                                            p.insert(3, 0)
                                        else:
                                            p.insert(2, 0)
                                        if tga_wl_id == 0:
                                            gpa_pa_org[i1].append(p)
                                        elif tga_wl_id == 1:
                                            gpa_pa_hum[i1].append(p)
                                        else:
                                            gpa_pa_hyd[i1].append(p)
        tga_wl_id += 1

    gpa_pa_org = [x for x in gpa_pa_org]
    gpa_pa_hum = [x for x in gpa_pa_hum]
    gpa_pa_hyd = [x for x in gpa_pa_hyd]
    gpa_pa = [[] for i in range(len(labels))]
    for i in range(len(gpa_pa_org)):
        for j in range(len(gpa_pa_org[i])):
            gpa_pa[i].append(gpa_pa_org[i][j])
        for j in range(len(gpa_pa_hum[i])):
            gpa_pa[i].append(gpa_pa_hum[i][j])
        for j in range(len(gpa_pa_hyd[i])):
            gpa_pa[i].append(gpa_pa_hyd[i][j])

    phases_org = {}
    phases_hum = {}
    phases_hyd = {}
    phases = {}
    for i in range(len(labels)):
        phases_org[labels[i]] = dict(zip(["mean", "std"], [np.mean(gpa_pa_org[i], axis=0).tolist(),
                                                           np.std(gpa_pa_org[i], axis=0).tolist()]))
        phases_hum[labels[i]] = dict(zip(["mean", "std"], [np.mean(gpa_pa_hum[i], axis=0).tolist(),
                                                           np.std(gpa_pa_hum[i], axis=0).tolist()]))
        phases_hyd[labels[i]] = dict(zip(["mean", "std"], [np.mean(gpa_pa_hyd[i], axis=0).tolist(),
                                                           np.std(gpa_pa_hyd[i], axis=0).tolist()]))
        phases[labels[i]] = dict(zip(["mean", "std"], [np.mean(gpa_pa[i], axis=0).tolist(),
                                                       np.std(gpa_pa[i], axis=0).tolist()]))
    return phases, phases_org, phases_hum, phases_hyd


def load_excel(filename):
    wb = openpyxl.load_workbook(filename, data_only=True)
    sh_labels = wb.worksheets[0]
    sh_wl = wb.worksheets[1]
    sh_wg = wb.worksheets[2]
    err = 2
    samples = {}
    i = 0
    for row in range(sh_labels.max_row):
        if sh_labels.cell(row=row + 2, column=1).value:
            samples[i] = {}
            samples[i]['wl'] = {'org': [], 'hum': [], 'hyd': []}
            samples[i]['wg'] = {'hum': [], 'hyd': []}
            if sh_labels.cell(row=row + 2, column=1).value == sh_wl.cell(row=row + 2, column=1).value \
                    and sh_labels.cell(row=row + 2, column=1).value == sh_wg.cell(row=row + 2, column=1).value:

                samples[i]['Label'] = sh_labels.cell(row=row + 2, column=1).value
                for col in range(3):
                    if sh_wl.cell(row=row + 2, column=col + 2).value:
                        err = 0
                        samples[i]['wl']['org'].append(sh_wl.cell(row=row + 2, column=col + 2).value)
                    if sh_wl.cell(row=row + 2, column=col + 5).value:
                        err = 0
                        samples[i]['wl']['hum'].append(sh_wl.cell(row=row + 2, column=col + 5).value)
                    if sh_wl.cell(row=row + 2, column=col + 8).value:
                        err = 0
                        samples[i]['wl']['hyd'].append(sh_wl.cell(row=row + 2, column=col + 8).value)

                    if sh_wg.cell(row=row + 2, column=col + 5).value:
                        err = 0
                        samples[i]['wg']['hum'].append(sh_wg.cell(row=row + 2, column=col + 5).value)
                    if sh_wg.cell(row=row + 2, column=col + 8).value:
                        err = 0
                        samples[i]['wg']['hyd'].append(sh_wg.cell(row=row + 2, column=col + 8).value)
            else:
                err = 1
            i += 1

    if err != 0:
        samples = {}

    return samples, err


def save_excel(filename, phases):
    """

        :param address: address at which the output excel file should be saved
        :param phases: a phases dictionary:
        labels: name of the analysis (like "TGA" or "GPA"). This will be the name of the sheet.
        values: phase calculations outputs of tga_solve or gpa_solve functions. These are dictionaries whose labels are
        sample labels and values are mean and std of each phase content.
        :return: None
        """

    wb = openpyxl.Workbook()
    counter = 0
    for analysis in phases:
        if not isListEmpty(phases[analysis]):
            if counter == 0:
                ws = wb.active
                ws.title = analysis
            else:
                ws = wb.create_sheet(analysis)

            p = phases[analysis]
            thin = openpyxl.styles.Side(border_style="thin")
            double = openpyxl.styles.Side(border_style="double")
            color = ['DDDDDD', 'DDDDDD', 'C7CEFF', 'C7CEFF', 'FFC7CE', 'FFC7CE', 'CEFFC7', 'CEFFC7', 'FFCEC7', 'FFCEC7']
            ws['A1'] = 'Labels'
            ws.merge_cells('B1:C1')
            ws['B1'] = 'DH'
            ws.merge_cells('D1:E1')
            ws['D1'] = 'HH'
            ws.merge_cells('F1:G1')
            ws['F1'] = 'AIII'
            ws.merge_cells('H1:I1')
            ws['H1'] = 'FM'
            ws.merge_cells('J1:K1')
            ws['J1'] = 'IN'
            for i in range(len(p)+2):
                for j in range(10):
                    c = ws.cell(row=1, column=j + 2)
                    c.fill = openpyxl.styles.PatternFill("solid", fgColor=color[j])
                    c.alignment = openpyxl.styles.Alignment(horizontal='center')
                    c = ws.cell(row=2, column=j + 2)
                    c.fill = openpyxl.styles.PatternFill("solid", fgColor=color[j])
                    c.alignment = openpyxl.styles.Alignment(horizontal='center')
                    c = ws.cell(row=i + 1, column=j + 2)
                    k = j % 2
                    if k == 0:
                        if i == 1:
                            ws.cell(row=i + 1, column=j + 2 + k, value='mean')
                        c.border = openpyxl.styles.Border(top=thin, left=double, right=thin, bottom=thin)
                        c.alignment = openpyxl.styles.Alignment(horizontal='center')
                    if k == 1:
                        if i == 1:
                            ws.cell(row=i + 1, column=j + 1 + k, value='std')
                        c.border = openpyxl.styles.Border(top=thin, left=thin, right=double, bottom=thin)
                        c.alignment = openpyxl.styles.Alignment(horizontal='center')
            for c in ws[1]:
                c.font = openpyxl.styles.Font(bold=True)
            length = 0
            i1 = 0
            for label in p:
                if label:
                    if len(str(label)) > length:
                        length = len(str(label))
                    c = ws.cell(row=3 + i1, column=1, value=str(label))
                    c.border = openpyxl.styles.Border(top=thin, bottom=thin)
                for i2 in range(5):
                    ws.cell(row=3 + i1, column=2 + i2 * 2, value=p[label]['mean'][i2])
                    ws.cell(row=3 + i1, column=3 + i2 * 2, value=p[label]['std'][i2])
                i1 += 1
            ws.column_dimensions["A"].width = length
            counter += 1
    wb.save(filename)
